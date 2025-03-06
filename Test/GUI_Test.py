import sys
from PyQt5.QtCore import Qt, QUrl, QStringListModel, QEvent
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QSizePolicy, QLineEdit,
    QFrame, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QComboBox,
    QCompleter, QFileDialog
)
from PyQt5.QtGui import QPalette, QColor, QDesktopServices
from PyQt5.QtWebEngineWidgets import QWebEngineView, QWebEnginePage
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class WebPage(QWebEnginePage):
    link_clicked = None

    def acceptNavigationRequest(self, url, _type, isMainFrame):
        if _type == QWebEnginePage.NavigationTypeLinkClicked and isMainFrame:
            if self.link_clicked:
                self.link_clicked(url)
            return False
        return super().acceptNavigationRequest(url, _type, isMainFrame)


class LinkPreviewer(QMainWindow):
    def __init__(self, excel_file_path):
        super().__init__()

        self.excel_file_path = excel_file_path

        self.setWindowTitle("Master Design Hub Pre Selection Review")
        self.setGeometry(50, 80, 1800, 950)

        # Set a dark color palette
        dark_palette = QPalette()
        dark_palette.setColor(QPalette.Window, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.WindowText, QColor(255, 255, 255))
        dark_palette.setColor(QPalette.Base, QColor(42, 42, 42))
        dark_palette.setColor(QPalette.AlternateBase, QColor(66, 66, 66))
        dark_palette.setColor(QPalette.ToolTipBase, QColor(255, 255, 255))
        dark_palette.setColor(QPalette.ToolTipText, QColor(255, 255, 255))
        dark_palette.setColor(QPalette.Text, QColor(255, 255, 255))
        dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.ButtonText, QColor(255, 255, 255))
        dark_palette.setColor(QPalette.BrightText, QColor(255, 0, 0))
        dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.HighlightedText, QColor(0, 0, 0))

        self.setPalette(dark_palette)

        self.central_widget = QFrame()
        self.setCentralWidget(self.central_widget)

        # Check if the provided file path is valid
        if not self.check_file_path():
            sys.exit()

        # Create a top layout for search box
        self.top_layout = QHBoxLayout()

        # Create a search box
        self.search_entry = QLineEdit()
        self.search_entry.setPlaceholderText("Enter Design Name to Search")
        self.search_entry.setStyleSheet("QLineEdit { color: white; background-color: #363636; }")
        self.top_layout.addWidget(self.search_entry)

        # Create a completer for auto-completion with partial matching
        self.completer_model = QStringListModel()
        self.completer = QCompleter()
        self.completer.setModel(self.completer_model)
        self.completer.setFilterMode(Qt.MatchContains)  # Set filter mode to partial matching
        self.search_entry.setCompleter(self.completer)

        # Create a search button
        self.search_button = QPushButton("Search")
        self.search_button.clicked.connect(self.search_design)
        self.search_button.setStyleSheet("QPushButton { color: white; background-color: #363636; }")
        self.top_layout.addWidget(self.search_button, alignment=Qt.AlignRight)

        # Add the top layout to the main layout
        self.central_layout = QVBoxLayout(self.central_widget)
        self.central_layout.addLayout(self.top_layout)

        self.title_label = QLabel()
        self.title_label.setStyleSheet("QLabel { color: white; }")
        self.subtitle_label = QLabel()
        self.subtitle_label.setStyleSheet("QLabel { color: white; }")
        self.central_layout.addWidget(self.title_label)
        self.central_layout.addWidget(self.subtitle_label)

        self.link_entry = QLineEdit()
        self.link_entry.returnPressed.connect(self.preview_link)
        self.link_entry.setStyleSheet("QLineEdit { color: white; background-color: #363636; }")
        self.central_layout.addWidget(self.link_entry)

        self.browser = QWebEngineView()
        self.browser.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.browser.setPage(WebPage(self.browser.page().profile(), self))
        WebPage.link_clicked = self.open_link

        self.central_layout.addWidget(self.browser)

        self.status_label = QLabel("Status:")
        self.status_label.setStyleSheet("QLabel { color: white; }")
        self.central_layout.addWidget(self.status_label)

        self.status_combo = QComboBox()
        self.status_combo.addItems(["-", "Approved", "Review", "Rejected"])
        self.status_combo.setStyleSheet("QComboBox { color: white; background-color: #363636; }")
        self.central_layout.addWidget(self.status_combo)

        self.status_resp_label = QLabel("Status Resp:")
        self.status_resp_label.setStyleSheet("QLabel { color: white; }")
        self.central_layout.addWidget(self.status_resp_label)

        self.status_resp_combo = QComboBox()
        self.status_resp_combo.addItems(["-", "Eduardo", "Tony", "Hector", "Alex", "Oscar"])
        self.status_resp_combo.setStyleSheet("QComboBox { color: white; background-color: #363636; }")
        self.central_layout.addWidget(self.status_resp_combo)

        self.to_be_reviewed_resp_label = QLabel("To-be-reviewed Resp:")
        self.to_be_reviewed_resp_label.setStyleSheet("QLabel { color: white; }")
        self.central_layout.addWidget(self.to_be_reviewed_resp_label)

        self.to_be_reviewed_resp_entry = QLineEdit()
        self.to_be_reviewed_resp_entry.setPlaceholderText("Enter To-be-reviewed Resp")
        self.to_be_reviewed_resp_entry.setStyleSheet("QLineEdit { color: white; background-color: #363636; }")
        self.central_layout.addWidget(self.to_be_reviewed_resp_entry)

        self.notes_label = QLabel("Notes:")
        self.notes_label.setStyleSheet("QLabel { color: white; }")
        self.central_layout.addWidget(self.notes_label)

        self.notes_entry = QLineEdit()
        self.notes_entry.setPlaceholderText("Enter Notes")
        self.notes_entry.setStyleSheet("QLineEdit { color: white; background-color: #363636; }")
        self.central_layout.addWidget(self.notes_entry)

        # Create a QHBoxLayout for the buttons
        self.buttons_layout = QHBoxLayout()

        # Add the Previous, Next, and Save buttons to the buttons layout
        self.prev_button = QPushButton("Previous")
        self.prev_button.clicked.connect(self.show_previous_and_save)
        self.prev_button.setStyleSheet("QPushButton { color: white; background-color: #363636; }")
        self.buttons_layout.addWidget(self.prev_button)

        self.next_button = QPushButton("Next")
        self.next_button.clicked.connect(self.show_next_and_save)
        self.next_button.setStyleSheet("QPushButton { color: white; background-color: #363636; }")
        self.buttons_layout.addWidget(self.next_button)

        self.save_button = QPushButton("Save")
        self.save_button.clicked.connect(self.save_changes)
        self.save_button.setStyleSheet("QPushButton { color: white; background-color: #363636; }")
        self.buttons_layout.addWidget(self.save_button)

        # Add the buttons layout to the central layout
        self.central_layout.addLayout(self.buttons_layout)

        # Load links from Excel file
        self.load_data_from_excel()

        # Initialize current_link_index
        self.current_link_index = 0

        # Show the first link and title/subtitle
        self.show_current_link()

        # Populate completer with Design Names
        design_names = self.links_df["Design Name"].astype(str).tolist()
        self.completer_model.setStringList(design_names)

    def check_file_path(self):
        try:
            pd.read_excel(self.excel_file_path, engine='openpyxl')
            return True
        except FileNotFoundError:
            print("File not found.")
            return False
        except Exception as e:
            print(f"Error: {e}")
            return False

    def search_design(self):
        design_name_to_search = self.search_entry.text().strip()
        if design_name_to_search:
            design_index = self.links_df.index[
                self.links_df['Design Name'].str.contains(design_name_to_search, case=False)].tolist()
            if design_index:
                self.current_link_index = design_index[0]
                self.show_current_link()

    def load_data_from_excel(self):
        # Load links from Excel file
        self.links_df = pd.read_excel(self.excel_file_path, engine='openpyxl')

    def preview_link(self):
        link = self.link_entry.text()
        if link:
            url = QUrl(link)
            self.browser.setUrl(url)

    def show_previous_and_save(self):
        self.save_changes()
        self.show_previous()

    def show_next_and_save(self):
        self.save_changes()
        self.show_next()

    def show_previous(self):
        if self.current_link_index > 0:
            self.current_link_index -= 1
            self.show_current_link()

    def show_next(self):
        if self.current_link_index < len(self.links_df) - 1:
            self.current_link_index += 1
            self.show_current_link()

    def show_current_link(self):
        # Reload data from Excel to get the latest changes
        self.load_data_from_excel()

        current_link = self.links_df.loc[self.current_link_index, "Link (Americas)"]
        self.link_entry.setText(current_link)
        url = QUrl(current_link)
        self.browser.setUrl(url)

        # Update title and subtitle labels
        design_name = self.links_df.loc[self.current_link_index, "Design Name"]
        mfg = self.links_df.loc[self.current_link_index, "MFG"]
        self.title_label.setText(f"Design Name: {design_name}")
        self.subtitle_label.setText(f"MFG: {mfg}")

        # Update Status, Status Resp, To-be-reviewed Resp, and Notes entries
        status = str(self.links_df.loc[self.current_link_index, "Status"])
        status_resp = str(self.links_df.loc[self.current_link_index, "Status Resp"])
        to_be_reviewed_resp = str(self.links_df.loc[self.current_link_index, "To-be-reviewed Resp"])

        # Handle NaN values in the "Notes" column
        notes = self.links_df.loc[self.current_link_index, "Notes"]
        notes = str(notes) if pd.notna(notes) else ""

        self.status_combo.setCurrentText(status)
        self.status_resp_combo.setCurrentText(status_resp)
        self.to_be_reviewed_resp_entry.setText(to_be_reviewed_resp)
        self.notes_entry.setText(notes)

    def save_changes(self):
        # Save changes back to the Excel file while preserving formatting
        current_link = self.links_df.loc[self.current_link_index, "Link (Americas)"]

        # Load existing Excel workbook
        book = load_workbook(self.excel_file_path)

        # Select the appropriate sheet
        sheet = book.active  # You may need to specify the sheet name if there are multiple sheets

        # Find the row corresponding to the current link
        row_number = None
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                if cell.value == current_link:
                    row_number = cell.row
                    break
            if row_number:
                break

        status_column = 'E'
        status_resp_column = 'F'
        to_be_reviewed_resp_column = 'G'
        notes_column = 'I'
        
        # Update the "Status," "Status Resp," "To-be-reviewed Resp," and "Notes" columns
        status_column_index = column_index_from_string(status_column)
        status_resp_column_index = column_index_from_string(status_resp_column)
        to_be_reviewed_resp_column_index = column_index_from_string(to_be_reviewed_resp_column)
        notes_column_index = column_index_from_string(notes_column)

        sheet.cell(row=row_number, column=status_column_index, value=self.status_combo.currentText())
        sheet.cell(row=row_number, column=status_resp_column_index, value=self.status_resp_combo.currentText())
        sheet.cell(row=row_number, column=to_be_reviewed_resp_column_index, value=self.to_be_reviewed_resp_entry.text())
        sheet.cell(row=row_number, column=notes_column_index, value=self.notes_entry.text())

        # Save changes back to the Excel file
        book.save(self.excel_file_path)

    def link_hovered(self, link):
        self.browser.setCursor(Qt.PointingHandCursor)

    def open_link(self, url):
        QDesktopServices.openUrl(url)


def get_excel_file_path():
    options = QFileDialog.Options()
    options |= QFileDialog.DontUseNativeDialog
    file_path, _ = QFileDialog.getOpenFileName(None, "Select Excel Master Design Hub Pre Selection Review File", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
    return file_path


if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    while True:
        excel_file_path = get_excel_file_path()
        if excel_file_path:
            break
        else:
            sys.exit()

    window = LinkPreviewer(excel_file_path)
    window.show()
    sys.exit(app.exec_())